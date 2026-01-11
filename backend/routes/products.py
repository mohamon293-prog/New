"""
Products Routes for Gamelo API
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import csv
import io

from models.schemas import ProductCreate, ProductResponse, CategoryCreate, CategoryResponse
from utils.database import db, fernet, get_current_user, get_admin_user

router = APIRouter(tags=["Products"])


@router.get("/categories", response_model=List[CategoryResponse])
async def get_categories():
    """Get all active categories"""
    categories = await db.categories.find(
        {"is_active": True}, {"_id": 0}
    ).sort("order", 1).to_list(100)
    return categories


@router.get("/products", response_model=List[ProductResponse])
async def get_products(
    category_id: Optional[str] = None,
    platform: Optional[str] = None,
    search: Optional[str] = None,
    featured: Optional[bool] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get products with optional filters"""
    query = {"is_active": True}
    
    if category_id:
        query["category_id"] = category_id
    if platform:
        query["platform"] = platform
    if featured:
        query["is_featured"] = True
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"name_en": {"$regex": search, "$options": "i"}}
        ]
    
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    # Get stock counts
    for product in products:
        stock = await db.codes.count_documents({
            "product_id": product["id"],
            "is_sold": False
        })
        product["stock_count"] = stock
        product["available_codes"] = stock
        
        # Get category name
        category = await db.categories.find_one({"id": product.get("category_id")})
        product["category_name"] = category.get("name") if category else None
    
    return products


@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str):
    """Get single product by ID or slug"""
    # Try to find by ID first, then by slug
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    
    if not product:
        # Try by slug
        product = await db.products.find_one({"slug": product_id}, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # Get stock count (available codes)
    stock = await db.codes.count_documents({
        "product_id": product_id,
        "is_sold": False
    })
    product["stock_count"] = stock
    product["available_codes"] = stock
    
    # Get category name
    category = await db.categories.find_one({"id": product.get("category_id")})
    product["category_name"] = category.get("name") if category else None
    
    return product


# Public stats endpoint
@router.get("/stats/public")
async def get_public_stats():
    """Get public statistics for homepage"""
    total_users = await db.users.count_documents({})
    total_orders = await db.orders.count_documents({})
    total_products = await db.products.count_documents({"is_active": True})
    
    return {
        "total_users": total_users,
        "total_orders": total_orders,
        "total_products": total_products,
        "satisfaction_rate": 100
    }


# Admin Products Endpoints
@router.get("/admin/products")
async def get_admin_products(
    skip: int = 0,
    limit: int = 50,
    admin: dict = Depends(get_admin_user)
):
    """Get all products with stock counts for admin"""
    products = await db.products.find({}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    for product in products:
        stock = await db.codes.count_documents({
            "product_id": product["id"],
            "is_sold": False
        })
        product["stock_count"] = stock
        
        category = await db.categories.find_one({"id": product.get("category_id")})
        product["category_name"] = category.get("name") if category else None
    
    return products


@router.post("/admin/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, admin: dict = Depends(get_admin_user)):
    """Create a new product"""
    product_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Generate slug if not provided
    slug = product.slug
    if not slug:
        import re
        slug = re.sub(r'[^\w\s-]', '', product.name.lower())
        slug = re.sub(r'[-\s]+', '-', slug).strip('-')
        slug = f"{slug}-{product_id[:8]}"
    
    # Process variants
    variants_data = None
    if product.has_variants and product.variants:
        variants_data = []
        for v in product.variants:
            variant_dict = v.model_dump() if hasattr(v, 'model_dump') else v.dict()
            if not variant_dict.get("id"):
                variant_dict["id"] = str(uuid.uuid4())
            variants_data.append(variant_dict)
    
    product_doc = {
        "id": product_id,
        **product.model_dump(),
        "slug": slug,
        "variants": variants_data,
        "stock_count": 0,
        "is_active": True,
        "rating": 0.0,
        "review_count": 0,
        "sold_count": 0,
        "created_at": now,
        "updated_at": now
    }
    
    await db.products.insert_one(product_doc)
    
    category = await db.categories.find_one({"id": product.category_id})
    product_doc["category_name"] = category.get("name") if category else None
    
    return product_doc


@router.put("/admin/products/{product_id}")
async def update_product(
    product_id: str,
    product: ProductCreate,
    admin: dict = Depends(get_admin_user)
):
    """Update a product"""
    existing = await db.products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Process variants
    variants_data = None
    if product.has_variants and product.variants:
        variants_data = []
        for v in product.variants:
            variant_dict = v.model_dump() if hasattr(v, 'model_dump') else v.dict()
            if not variant_dict.get("id"):
                variant_dict["id"] = str(uuid.uuid4())
            variants_data.append(variant_dict)
    
    update_data = {
        **product.model_dump(),
        "variants": variants_data,
        "updated_at": now
    }
    
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    
    return {"message": "تم تحديث المنتج"}


@router.delete("/admin/products/{product_id}")
async def delete_product(product_id: str, permanent: bool = False, admin: dict = Depends(get_admin_user)):
    """Delete a product (soft or permanent)"""
    if permanent:
        # Permanent delete
        result = await db.products.delete_one({"id": product_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="المنتج غير موجود")
        # Also delete associated codes
        await db.codes.delete_many({"product_id": product_id})
        return {"message": "تم حذف المنتج نهائياً"}
    else:
        # Soft delete
        result = await db.products.update_one(
            {"id": product_id},
            {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="المنتج غير موجود")
        return {"message": "تم تعطيل المنتج"}


@router.get("/admin/products/{product_id}/codes")
async def get_product_codes(
    product_id: str,
    status: Optional[str] = None,
    admin: dict = Depends(get_admin_user)
):
    """Get product codes for admin"""
    query = {"product_id": product_id}
    if status == "available":
        query["is_sold"] = False
    elif status == "sold":
        query["is_sold"] = True
    
    codes = await db.codes.find(query, {"_id": 0}).to_list(500)
    
    # Decrypt codes for display
    for code in codes:
        try:
            code["code_value"] = fernet.decrypt(code["code_value"].encode()).decode()
        except:
            pass
    
    return {
        "codes": codes,
        "total": len(codes),
        "available": sum(1 for c in codes if not c.get("is_sold")),
        "sold": sum(1 for c in codes if c.get("is_sold"))
    }


@router.post("/admin/products/{product_id}/codes/upload")
async def upload_product_codes(
    product_id: str,
    file: UploadFile = File(...),
    admin: dict = Depends(get_admin_user)
):
    """Upload product codes from CSV file"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    content = await file.read()
    try:
        text = content.decode('utf-8')
    except:
        text = content.decode('utf-8-sig')
    
    reader = csv.reader(io.StringIO(text))
    codes_added = 0
    now = datetime.now(timezone.utc).isoformat()
    
    for row in reader:
        if not row:
            continue
        
        code_value = row[0].strip()
        if not code_value:
            continue
        
        # Check for duplicate
        existing = await db.codes.find_one({
            "product_id": product_id,
            "code_value": fernet.encrypt(code_value.encode()).decode()
        })
        
        if existing:
            continue
        
        code_doc = {
            "id": str(uuid.uuid4()),
            "product_id": product_id,
            "code_value": fernet.encrypt(code_value.encode()).decode(),
            "is_sold": False,
            "created_at": now
        }
        
        await db.codes.insert_one(code_doc)
        codes_added += 1
    
    # Update product stock count
    await db.products.update_one(
        {"id": product_id},
        {"$inc": {"stock_count": codes_added}}
    )
    
    return {"added": codes_added, "message": f"تم إضافة {codes_added} كود بنجاح"}


@router.post("/admin/products/{product_id}/codes/add")
async def add_product_codes(
    product_id: str,
    codes: List[str],
    admin: dict = Depends(get_admin_user)
):
    """Add product codes manually"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    now = datetime.now(timezone.utc).isoformat()
    codes_added = 0
    
    for code_value in codes:
        code_value = code_value.strip()
        if not code_value:
            continue
        
        encrypted = fernet.encrypt(code_value.encode()).decode()
        existing = await db.codes.find_one({
            "product_id": product_id,
            "code_value": encrypted
        })
        
        if existing:
            continue
        
        code_doc = {
            "id": str(uuid.uuid4()),
            "product_id": product_id,
            "code_value": encrypted,
            "is_sold": False,
            "created_at": now
        }
        
        await db.codes.insert_one(code_doc)
        codes_added += 1
    
    # Update product stock count
    await db.products.update_one(
        {"id": product_id},
        {"$inc": {"stock_count": codes_added}}
    )
    
    return {"added": codes_added, "message": f"تم إضافة {codes_added} كود بنجاح"}


# Admin Categories
@router.get("/admin/categories")
async def get_admin_categories(admin: dict = Depends(get_admin_user)):
    """Get all categories for admin"""
    categories = await db.categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return categories


@router.post("/admin/categories", response_model=CategoryResponse)
async def create_category(category: CategoryCreate, admin: dict = Depends(get_admin_user)):
    """Create a new category"""
    existing = await db.categories.find_one({"slug": category.slug})
    if existing:
        raise HTTPException(status_code=400, detail="القسم موجود مسبقاً")
    
    category_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    category_doc = {
        "id": category_id,
        **category.model_dump(),
        "is_active": True,
        "created_at": now,
        "updated_at": now
    }
    
    await db.categories.insert_one(category_doc)
    return category_doc


@router.put("/admin/categories/{category_id}")
async def update_category(
    category_id: str,
    category: CategoryCreate,
    admin: dict = Depends(get_admin_user)
):
    """Update a category"""
    result = await db.categories.update_one(
        {"id": category_id},
        {"$set": {**category.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    return {"message": "تم تحديث القسم"}


@router.delete("/admin/categories/{category_id}")
async def delete_category(category_id: str, admin: dict = Depends(get_admin_user)):
    """Soft delete a category"""
    result = await db.categories.update_one(
        {"id": category_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="القسم غير موجود")
    
    return {"message": "تم حذف القسم"}



# Excel Import
@router.post("/admin/products/import/excel")
async def import_products_excel(
    file: UploadFile = File(...),
    admin: dict = Depends(get_admin_user)
):
    """Import products from Excel/CSV file"""
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. استخدم CSV أو Excel")
    
    try:
        content = await file.read()
        
        # Handle CSV
        if file.filename.endswith('.csv'):
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            # For Excel, we need openpyxl
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(content))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        rows.append(dict(zip(headers, row)))
            except ImportError:
                raise HTTPException(status_code=400, detail="لقراءة ملفات Excel، يرجى رفع ملف CSV بدلاً منه")
        
        products_added = 0
        codes_added = 0
        errors = []
        now = datetime.now(timezone.utc).isoformat()
        
        for idx, row in enumerate(rows, start=2):
            try:
                # Required fields
                name = row.get('name') or row.get('الاسم')
                if not name:
                    errors.append(f"صف {idx}: الاسم مطلوب")
                    continue
                
                category_id = row.get('category') or row.get('category_id') or row.get('القسم') or 'other'
                price_jod = float(row.get('price_jod') or row.get('السعر') or 0)
                price_usd = float(row.get('price_usd') or row.get('السعر_دولار') or price_jod * 1.41)
                
                # Generate slug
                slug = row.get('slug') or name.lower().replace(' ', '-').replace('$', '').replace('/', '-')
                slug = f"{slug}-{uuid.uuid4().hex[:6]}"
                
                product_id = str(uuid.uuid4())
                product_type = row.get('type') or row.get('product_type') or row.get('النوع') or 'digital_code'
                
                product_doc = {
                    "id": product_id,
                    "name": name,
                    "name_en": row.get('name_en') or row.get('الاسم_انجليزي') or name,
                    "slug": slug,
                    "description": row.get('description') or row.get('الوصف') or f"منتج {name}",
                    "description_en": row.get('description_en') or None,
                    "category_id": category_id,
                    "price_jod": price_jod,
                    "price_usd": price_usd,
                    "original_price_jod": float(row.get('original_price_jod') or 0) or None,
                    "original_price_usd": float(row.get('original_price_usd') or 0) or None,
                    "image_url": row.get('image_url') or row.get('الصورة') or f"https://placehold.co/400x400/1a1a2e/ffffff?text={name[:10]}",
                    "platform": row.get('platform') or row.get('المنصة') or category_id,
                    "region": row.get('region') or row.get('المنطقة') or "عالمي",
                    "is_active": True,
                    "is_featured": str(row.get('is_featured') or row.get('مميز') or '').lower() in ('true', '1', 'yes', 'نعم'),
                    "product_type": product_type,
                    "has_variants": False,
                    "variants": None,
                    "requires_email": product_type == 'existing_account',
                    "requires_password": product_type == 'existing_account',
                    "requires_phone": product_type == 'new_account',
                    "delivery_instructions": row.get('delivery_instructions') or row.get('تعليمات') or None,
                    "rating": 5.0,
                    "review_count": 0,
                    "sold_count": 0,
                    "stock_count": 0,
                    "created_at": now,
                    "updated_at": now
                }
                
                await db.products.insert_one(product_doc)
                products_added += 1
                
                # Add codes if provided
                codes_str = row.get('codes') or row.get('الأكواد') or ''
                if codes_str and product_type == 'digital_code':
                    codes_list = str(codes_str).split('|')
                    for code_value in codes_list:
                        code_value = code_value.strip()
                        if not code_value:
                            continue
                        
                        encrypted = fernet.encrypt(code_value.encode()).decode()
                        code_doc = {
                            "id": str(uuid.uuid4()),
                            "product_id": product_id,
                            "code_value": encrypted,
                            "is_sold": False,
                            "created_at": now
                        }
                        await db.codes.insert_one(code_doc)
                        codes_added += 1
                    
                    # Update stock count
                    await db.products.update_one(
                        {"id": product_id},
                        {"$set": {"stock_count": len([c for c in codes_list if c.strip()])}}
                    )
                
            except Exception as e:
                errors.append(f"صف {idx}: {str(e)}")
        
        return {
            "success": True,
            "products_added": products_added,
            "codes_added": codes_added,
            "errors": errors[:10] if errors else [],
            "message": f"تم استيراد {products_added} منتج و {codes_added} كود"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")


@router.get("/admin/products/import/template")
async def get_import_template(admin: dict = Depends(get_admin_user)):
    """Get CSV template for product import"""
    return {
        "headers": ["name", "name_en", "category", "type", "price_jod", "price_usd", "image_url", "region", "codes", "description"],
        "example_row": ["بطاقة ستيم $10", "Steam $10", "steam", "digital_code", "7.5", "10", "https://example.com/img.jpg", "عالمي", "CODE1|CODE2|CODE3", "وصف المنتج"],
        "categories": ["playstation", "xbox", "steam", "nintendo", "mobile", "giftcards"],
        "types": ["digital_code", "existing_account", "new_account"]
    }

